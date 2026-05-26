"""Tests del módulo `apps.web.instruments_abm`.

Foco: transactionality + atomic save (Batch 1 del audit fix).
"""

import os
import shutil
from datetime import datetime

import openpyxl
import pytest

from apps.web.instruments_abm import (
    _atomic_save_workbook,
    _parse_cashflows,
    delete_instrument,
    get_instrument,
    list_instruments,
    save_cashflows,
    save_instrument,
)


@pytest.fixture
def tmp_master(tmp_path):
    """Copia el master real a un tmpfile aislado por test."""
    src = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "data", "instruments_master.xlsx",
    )
    dst = tmp_path / "master_test.xlsx"
    shutil.copy2(src, dst)
    return str(dst)


# --------------------------------------------------------------------------- #
# _parse_cashflows — validación de input del frontend
# --------------------------------------------------------------------------- #

class TestParseCashflows:
    def test_valid_input(self):
        rows = [
            {"date": "2026-07-09", "amortization": 8.0, "interest": 0.27},
            {"date": "2027-01-09", "amortization": 8.0, "interest": 0.24},
        ]
        parsed = _parse_cashflows(rows)
        assert len(parsed) == 2
        assert parsed[0][0].isoformat() == "2026-07-09"
        assert parsed[0][1] == 8.0
        assert parsed[0][2] == 0.27

    def test_sorts_by_date(self):
        rows = [
            {"date": "2027-01-09", "amortization": 1.0, "interest": 0.1},
            {"date": "2026-07-09", "amortization": 2.0, "interest": 0.2},
        ]
        parsed = _parse_cashflows(rows)
        assert parsed[0][0].isoformat() == "2026-07-09"
        assert parsed[1][0].isoformat() == "2027-01-09"

    def test_invalid_date_raises(self):
        with pytest.raises(ValueError, match="invalid date"):
            _parse_cashflows([{"date": "not-a-date", "amortization": 1, "interest": 0}])

    def test_empty_date_skipped(self):
        rows = [
            {"date": "", "amortization": 1, "interest": 0},
            {"date": "2026-07-09", "amortization": 2, "interest": 0},
        ]
        parsed = _parse_cashflows(rows)
        assert len(parsed) == 1

    def test_missing_amort_interest_default_zero(self):
        parsed = _parse_cashflows([{"date": "2026-07-09"}])
        assert parsed[0][1] == 0.0
        assert parsed[0][2] == 0.0


# --------------------------------------------------------------------------- #
# Atomic save: el .tmp + os.replace pattern
# --------------------------------------------------------------------------- #

class TestAtomicSave:
    def test_atomic_save_writes_then_replaces(self, tmp_path):
        path = str(tmp_path / "test.xlsx")
        wb = openpyxl.Workbook()
        wb.active.title = "Test"
        wb.active["A1"] = "hello"
        # Pre-existing file (simular update)
        wb.save(path)
        # Mutamos y re-guardamos atomic
        wb2 = openpyxl.load_workbook(path)
        wb2.active["A2"] = "world"
        _atomic_save_workbook(wb2, path)
        # El .tmp NO debe quedar después
        assert not os.path.exists(path + ".tmp")
        # El destino contiene la mutación
        wb3 = openpyxl.load_workbook(path, read_only=True)
        assert wb3["Test"]["A2"].value == "world"
        wb3.close()


# --------------------------------------------------------------------------- #
# save_instrument transaccional (row + cashflows en un solo wb handle)
# --------------------------------------------------------------------------- #

class TestSaveInstrumentTransactional:
    def test_save_row_only_no_cashflows(self, tmp_master):
        """Sin cashflows kwarg, save_instrument solo escribe la fila."""
        result = save_instrument(
            tmp_master,
            "Soberanos",
            {
                "ticker": "TESTROW",
                "short_name": "TESTROW",
                "tipo": "BONAR",
                "fecha_emision": "2025-01-01",
                "fecha_vencimiento": "2030-01-01",
                "cupon anual %": "1.5",
                "frecuencia pagos": "2",
            },
        )
        assert result["action"] == "created"
        assert result["ticker"] == "TESTROW"
        assert "cashflows" not in result

    def test_save_with_cashflows_atomic(self, tmp_master):
        """Con cashflows, ambas escrituras pasan en un solo wb save.
        Verificamos que ambas mutaciones persisten."""
        cashflows = [
            {"date": "2026-01-01", "amortization": 50.0, "interest": 1.0},
            {"date": "2027-01-01", "amortization": 50.0, "interest": 0.5},
        ]
        result = save_instrument(
            tmp_master,
            "Soberanos",
            {
                "ticker": "TXN1",
                "short_name": "TXN1",
                "tipo": "BONAR",
                "fecha_emision": "2025-01-01",
                "fecha_vencimiento": "2027-01-01",
                "cupon anual %": "2.0",
                "frecuencia pagos": "2",
            },
            cashflows=cashflows,
        )
        assert result["cashflows"] == 2

        # Verificar que get_instrument lee tanto row como cashflows
        loaded = get_instrument(tmp_master, "TXN1")
        assert loaded is not None
        assert loaded["fields"]["tipo"] == "BONAR"
        assert len(loaded["cashflows"]) == 2
        assert loaded["cashflows_source"] == "sheet"

    def test_save_invalid_cashflow_aborts_everything(self, tmp_master):
        """Si los cashflows tienen una fecha inválida, ni la row ni los
        cashflows deben persistir (validación pre-write)."""
        with pytest.raises(ValueError, match="invalid date"):
            save_instrument(
                tmp_master,
                "Soberanos",
                {
                    "ticker": "BADCF",
                    "short_name": "BADCF",
                    "tipo": "BONAR",
                    "fecha_emision": "2025-01-01",
                    "fecha_vencimiento": "2030-01-01",
                    "cupon anual %": "1.0",
                },
                cashflows=[
                    {"date": "GARBAGE", "amortization": 100, "interest": 0},
                ],
            )
        # Verificar que la row NO se creó
        assert get_instrument(tmp_master, "BADCF") is None

    def test_save_updates_existing(self, tmp_master):
        """Editar un ticker existente preserva otros campos no pasados."""
        # Primer save
        save_instrument(tmp_master, "Soberanos", {
            "ticker": "UPDT1", "short_name": "UPDT1", "tipo": "BONAR",
            "fecha_emision": "2025-01-01", "fecha_vencimiento": "2030-01-01",
            "cupon anual %": "1.0", "frecuencia pagos": "2",
        })
        # Update
        result = save_instrument(tmp_master, "Soberanos", {
            "ticker": "UPDT1", "short_name": "UPDT1",
            "tipo": "BONAR",  # mismo
            "fecha_emision": "2025-01-01",
            "fecha_vencimiento": "2030-01-01",
            "cupon anual %": "2.5",  # cambiado
            "frecuencia pagos": "2",
        })
        assert result["action"] == "updated"
        loaded = get_instrument(tmp_master, "UPDT1")
        assert str(loaded["fields"]["cupon anual %"]) in ("2.5", "2,5")


# --------------------------------------------------------------------------- #
# save_cashflows standalone
# --------------------------------------------------------------------------- #

class TestSaveCashflows:
    def test_replace_existing_cashflows(self, tmp_master):
        """save_cashflows reemplaza TODAS las filas del ticker (delete + insert)."""
        # AL30D tiene 9 cashflows en el master. Las reemplazamos por 2.
        result = save_cashflows(tmp_master, "AL30D", [
            {"date": "2030-01-01", "amortization": 100, "interest": 5},
        ])
        assert result["count"] == 1
        loaded = get_instrument(tmp_master, "AL30D")
        assert loaded["cashflows_source"] == "sheet"
        assert len(loaded["cashflows"]) == 1
        assert loaded["cashflows"][0]["amortization"] == 100

    def test_empty_cashflows_clears(self, tmp_master):
        """Pasar [] borra todas las cashflows del ticker."""
        save_cashflows(tmp_master, "AL30D", [])
        loaded = get_instrument(tmp_master, "AL30D")
        # Sin cashflows explícitos, fall-back a synth
        assert loaded["cashflows_source"] in ("synth", "empty")


# --------------------------------------------------------------------------- #
# delete_instrument
# --------------------------------------------------------------------------- #

class TestDeleteInstrument:
    def test_delete_removes_row_and_cashflows(self, tmp_master):
        # Crear ticker con cashflows
        save_instrument(tmp_master, "Soberanos", {
            "ticker": "DELME", "short_name": "DELME", "tipo": "BONAR",
            "fecha_emision": "2025-01-01", "fecha_vencimiento": "2030-01-01",
            "cupon anual %": "1.0",
        }, cashflows=[{"date": "2030-01-01", "amortization": 100, "interest": 0}])
        assert get_instrument(tmp_master, "DELME") is not None

        # Borrar
        result = delete_instrument(tmp_master, "DELME")
        assert result["action"] == "deleted"
        assert result["ticker"] == "DELME"
        assert get_instrument(tmp_master, "DELME") is None

    def test_delete_nonexistent_returns_not_found(self, tmp_master):
        result = delete_instrument(tmp_master, "NEVEREXISTED")
        assert result["action"] == "not_found"


# --------------------------------------------------------------------------- #
# list_instruments
# --------------------------------------------------------------------------- #

class TestListInstruments:
    def test_returns_known_tickers(self, tmp_master):
        items = list_instruments(tmp_master)
        tickers = {it["ticker"] for it in items}
        assert "AL30D" in tickers
        assert "S29Y6" in tickers

    def test_returns_sheet_per_ticker(self, tmp_master):
        items = list_instruments(tmp_master)
        sheet_by_ticker = {it["ticker"]: it["sheet"] for it in items}
        assert sheet_by_ticker.get("AL30D") == "Soberanos"
        assert sheet_by_ticker.get("S29Y6") == "Tasa_Fija"
