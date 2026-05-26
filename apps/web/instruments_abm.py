"""ABM (Alta/Baja/Modificación) backend for instruments_master.xlsx.

Provides JSON-friendly CRUD operations over the per-sheet rows of the master
Excel. Each sheet has its own schema; this module exposes them as a single
typed contract that the frontend can introspect.

Thread safety: a single module-level RLock serializes all writes. The Excel
file is small enough (<100 rows per sheet) that load → mutate → save in-place
is acceptable; pandas would lose formatting so we use openpyxl directly.
"""

from __future__ import annotations

import logging
import os
import threading
from datetime import date, datetime
from typing import Any, Dict, List, Optional

import openpyxl

logger = logging.getLogger(__name__)

# Lock global compartido por TODAS las operaciones I/O sobre el master.
# El loader del repo (ExcelInstrumentsRepository._load) también lo toma, así
# que un write nunca coincide con un read en progreso.
_LOCK = threading.RLock()


# ------------------------------------------------------------------------- #
# Atomic save helpers
# ------------------------------------------------------------------------- #

def _atomic_save_workbook(wb, xlsx_path: str) -> None:
    """Guardar el workbook a .tmp y luego os.replace al destino.

    Atomic en POSIX y Windows (NTFS) mientras src+dst estén en el mismo FS.
    Sin esto, un crash del proceso durante wb.save() dejaba el master
    corrupto/truncado y el siguiente boot fallaba al cargar instrumentos.
    """
    tmp = xlsx_path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, xlsx_path)

# Sheets that hold actual instruments (skip Cashflows tabs + metadata).
_INSTRUMENT_SHEETS = ("Soberanos", "Tasa_Fija", "CER", "Dolar_Linked", "TAMAR")

# Hoja Cashflows: una fila por (ticker, fecha) con amortizacion + cupon.
# El engine la consulta primero y si no hay rows para un ticker usa el synth
# desde los parametros del row (cupon, frecuencia, amort_inicio, etc.).
_CASHFLOWS_SHEET = "Cashflows"
_CASHFLOWS_COLS = ("ticker", "fecha_pago", "amortizacion", "cupon_interes")


def _synth_cashflows_for_fields(fields: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Genera cashflows on-the-fly desde un dict de fields del form ABM.

    Delegate al módulo puro `core.domain.cashflow_synth.synth_cashflows`.
    Antes accedíamos al método privado del Repository via __new__ hack —
    ahora es una import limpia.
    """
    try:
        from core.domain.cashflow_synth import synth_cashflows
        normalized = {str(k).lower().strip(): v for k, v in fields.items()}
        cfs = synth_cashflows(normalized)
        return [
            {
                "date": cf.date.isoformat() if hasattr(cf.date, "isoformat") else str(cf.date),
                "amortization": float(cf.amortization),
                "interest": float(cf.interest),
            }
            for cf in cfs
        ]
    except (ValueError, KeyError, TypeError, AttributeError) as e:
        # Catch narrow: errores de input mal-formado. Bugs reales (ImportError,
        # NameError) deben propagar para que se vean en el log con traceback.
        logger.warning(f"synth_cashflows_for_fields invalid input: {e}")
        return []


def get_cashflows(xlsx_path: str, ticker: str, fallback_fields: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Return {"rows": [...], "source": "sheet"|"synth"|"empty"} for `ticker`.

    Lee la hoja Cashflows primero. Si vacia y `fallback_fields` provisto,
    sintetiza desde esos params (con el mismo synth del engine).
    """
    ticker_u = ticker.upper().strip()
    rows: List[Dict[str, Any]] = []
    with _LOCK:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            if _CASHFLOWS_SHEET in wb.sheetnames:
                ws = wb[_CASHFLOWS_SHEET]
                hdr = None
                for r in ws.iter_rows(values_only=True):
                    if hdr is None:
                        hdr = [str(c).lower().strip() if c else "" for c in r]
                        continue
                    if not r or r[0] is None:
                        continue
                    if str(r[0]).strip().upper() != ticker_u:
                        continue
                    rec = dict(zip(hdr, r))
                    d = rec.get("fecha_pago")
                    if hasattr(d, "isoformat"):
                        d_str = d.isoformat()[:10]
                    else:
                        d_str = str(d) if d is not None else ""
                    rows.append({
                        "date": d_str,
                        "amortization": float(rec.get("amortizacion") or 0),
                        "interest": float(rec.get("cupon_interes") or 0),
                    })
        finally:
            wb.close()
    if rows:
        rows.sort(key=lambda x: x["date"])
        return {"rows": rows, "source": "sheet"}
    if fallback_fields:
        synth = _synth_cashflows_for_fields(fallback_fields)
        if synth:
            return {"rows": synth, "source": "synth"}
    return {"rows": [], "source": "empty"}


def _parse_cashflows(cashflows: List[Dict[str, Any]]) -> List[tuple]:
    """Validar + parsear lista del frontend a tuples (date, amort, interest)
    ordenadas por fecha. Aislado para reuso entre save_cashflows() standalone
    y save_instrument() transaccional."""
    from datetime import datetime
    parsed: List[tuple] = []
    for i, cf in enumerate(cashflows):
        d_raw = (cf.get("date") or "").strip()
        if not d_raw:
            continue
        try:
            d = datetime.strptime(d_raw[:10], "%Y-%m-%d").date()
        except ValueError:
            raise ValueError(f"cashflow #{i+1}: invalid date {d_raw!r} (expected YYYY-MM-DD)")
        amort = float(cf.get("amortization") or 0)
        interest = float(cf.get("interest") or 0)
        parsed.append((d, amort, interest))
    parsed.sort(key=lambda x: x[0])
    return parsed


def _write_cashflows_on_wb(wb, ticker_u: str, parsed: List[tuple]) -> None:
    """Mutar el workbook (en memoria) reemplazando todas las filas Cashflows
    del ticker. NO guarda — el caller decide cuándo. Permite combinar con
    otras escrituras en una única transacción atómica."""
    if _CASHFLOWS_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(_CASHFLOWS_SHEET)
        ws.append(list(_CASHFLOWS_COLS))
    else:
        ws = wb[_CASHFLOWS_SHEET]

    # Borrar rows existentes del ticker (de mayor a menor para no corromper
    # índices durante el delete).
    to_delete = [r for r in range(2, ws.max_row + 1)
                 if (v := ws.cell(row=r, column=1).value) and str(v).strip().upper() == ticker_u]
    for r in sorted(to_delete, reverse=True):
        ws.delete_rows(r, 1)

    for d, amort, interest in parsed:
        ws.append([ticker_u, d.isoformat(), amort, interest])


def save_cashflows(xlsx_path: str, ticker: str, cashflows: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Reemplaza TODAS las filas de Cashflows para `ticker` (atomic).

    Para escribir cashflows en la MISMA transacción que el row de instrument,
    usar save_instrument(..., cashflows=...) — comparte el wb handle.
    """
    ticker_u = ticker.upper().strip()
    if not ticker_u:
        raise ValueError("ticker is required")
    parsed = _parse_cashflows(cashflows)

    with _LOCK:
        wb = openpyxl.load_workbook(xlsx_path)
        try:
            _write_cashflows_on_wb(wb, ticker_u, parsed)
            _atomic_save_workbook(wb, xlsx_path)
            logger.info(f"ABM: saved {len(parsed)} cashflows for {ticker_u}")
            return {"ticker": ticker_u, "count": len(parsed)}
        finally:
            wb.close()

# Per-sheet field metadata used by the frontend to render the right input
# widget. `key` matches the Excel column header (lowercased + stripped).
_BASE_CALCULO_OPTIONS = ["", "ACT/365.25", "ACT/365", "30/360", "ACT/ACT"]
_TIPO_AMORT_OPTIONS  = ["", "bullet", "amortizing"]

SHEET_SCHEMAS: Dict[str, Dict[str, Any]] = {
    "Soberanos": {
        "label": "Soberanos (BONAR / GLOBAL / BOPREAL)",
        "fields": [
            {"key": "ticker",          "label": "Ticker",               "type": "text",   "required": True,
             "help": "Símbolo en Data912, ej. AL30D"},
            {"key": "short_name",      "label": "Short name",           "type": "text",   "required": True},
            {"key": "tipo",            "label": "Tipo",                 "type": "select", "required": True,
             "options": ["BONAR", "GLOBAL", "BOPREAL"]},
            {"key": "fecha_emision",   "label": "Fecha emisión",        "type": "date",   "required": True,
             "help": "Origen del schedule de cupones"},
            {"key": "fecha_vencimiento","label": "Vencimiento",         "type": "date",   "required": True},
            {"key": "cupon anual %",   "label": "Cupón anual %",        "type": "text",   "required": True,
             "help": "Ej. 1.00 (decimal). Step-up: '2024-12-31:0.63;2027-12-31:1.18'"},
            {"key": "frecuencia pagos","label": "Frecuencia pagos /año","type": "number", "step": "1",
             "help": "2 = semestral (default AR), 4 = trimestral"},
            {"key": "base calculo",    "label": "Base cálculo",         "type": "select",
             "options": _BASE_CALCULO_OPTIONS,
             "help": "Convención day-count para TIR y duración. BOPREAL: 30/360"},
            {"key": "tipo amortizacion","label": "Tipo amortización",   "type": "select",
             "options": _TIPO_AMORT_OPTIONS},
            {"key": "amort inicio",    "label": "Amort inicio",         "type": "date",
             "help": "Primera fecha de amortización (solo amortizing)"},
            {"key": "amort cantidad",  "label": "Cant. cuotas amort.",  "type": "number", "step": "1",
             "help": "Ej. 13 cuotas semestrales → 1 cashflow por cuota"},
            {"key": "capital factor",  "label": "Capital factor",       "type": "number", "step": "0.0001",
             "help": "Solo bonos reestructurados (DICP/CUAP). Dejar vacío para el resto"},
        ],
    },
    "Tasa_Fija": {
        "label": "Tasa Fija (LECAP / BONCAP / BONOFIJA)",
        "fields": [
            {"key": "ticker",           "label": "Ticker",                "type": "text",   "required": True},
            {"key": "clase",            "label": "Clase",                 "type": "select", "required": True,
             "options": ["", "LECAP", "BONCAP", "BONOFIJA"]},
            # Comunes a todas las clases — solo aparecen una vez elegida la clase
            {"key": "fecha_emision",    "label": "Fecha emisión",         "type": "date",   "required": True,
             "help": "Para LECAP/BONCAP determina la capitalización",
             "classes": ["LECAP", "BONCAP", "BONOFIJA"]},
            {"key": "fecha_pago",       "label": "Vencimiento",           "type": "date",   "required": True,
             "classes": ["LECAP", "BONCAP", "BONOFIJA"]},
            # Solo LECAP / BONCAP
            {"key": "tem_licit",        "label": "TEM licitación",        "type": "number", "step": "0.0001",
             "help": "Decimal: 0.021 = 2.1%", "classes": ["LECAP", "BONCAP"]},
            # Solo BONOFIJA
            {"key": "cupon anual %",    "label": "Cupón anual %",         "type": "text",
             "help": "Ej. 2.50 (porcentual anual)", "classes": ["BONOFIJA"]},
            {"key": "frecuencia pagos", "label": "Frecuencia pagos /año", "type": "number", "step": "1",
             "help": "2 = semestral (default)", "classes": ["BONOFIJA"]},
            # Común — aparece después de elegir clase
            {"key": "base calculo",     "label": "Base cálculo",          "type": "select",
             "options": _BASE_CALCULO_OPTIONS,
             "help": "LECAP/BONCAP: 30/360 (Sec. Finanzas)",
             "classes": ["LECAP", "BONCAP", "BONOFIJA"]},
            # Solo BONOFIJA: amortización
            {"key": "tipo amortizacion","label": "Tipo amortización",     "type": "select",
             "options": _TIPO_AMORT_OPTIONS, "classes": ["BONOFIJA"]},
            {"key": "amort inicio",     "label": "Amort inicio",          "type": "date",
             "help": "Primera fecha de amortización",
             "classes": ["BONOFIJA"], "show_if_amort": True},
            {"key": "amort cantidad",   "label": "Cant. cuotas amort.",   "type": "number", "step": "1",
             "classes": ["BONOFIJA"], "show_if_amort": True},
        ],
    },
    "CER": {
        "label": "CER (LECER / BONCER / BONCER ZC / CON CUPON / STEP-UP)",
        "fields": [
            {"key": "ticker",          "label": "Ticker",               "type": "text",   "required": True},
            {"key": "tipo",            "label": "Tipo",                 "type": "select", "required": True,
             "options": ["LECER", "BONCER", "BONCER ZC", "CON CUPON", "STEP-UP"]},
            {"key": "fecha emision",   "label": "Fecha emisión",        "type": "date",   "required": True},
            {"key": "fecha vencimiento","label": "Vencimiento",         "type": "date",   "required": True},
            {"key": "cupon anual %",   "label": "Cupón anual %",        "type": "text",
             "help": "Ej. 5.83. Step-up: '2024-12-31:0.63;2027-12-31:1.18'"},
            {"key": "frecuencia pagos","label": "Frecuencia pagos /año","type": "number", "step": "1",
             "help": "2 = semestral, 4 = trimestral"},
            {"key": "base calculo",    "label": "Base cálculo",         "type": "select",
             "options": _BASE_CALCULO_OPTIONS},
            {"key": "tipo amortizacion","label": "Tipo amortización",   "type": "select",
             "options": _TIPO_AMORT_OPTIONS},
            {"key": "amort inicio",    "label": "Amort inicio",         "type": "date",
             "help": "Solo amortizing"},
            {"key": "amort cantidad",  "label": "Cant. cuotas amort.",  "type": "number", "step": "1"},
            {"key": "capital factor",  "label": "Capital factor",       "type": "number", "step": "0.0001",
             "help": "Capitalización inicial, ej. 1.27 para DICP"},
            {"key": "cer emision",     "label": "CER base (10h pre-emisión)", "type": "number",
             "step": "0.000001", "help": "Crítico: CER 10 días hábiles antes de emisión"},
            {"key": "categoria",       "label": "Categoría",            "type": "text",
             "help": "Ej. 'BONCERES CERO CUPON'"},
        ],
    },
    "Dolar_Linked": {
        "label": "Dolar Linked",
        "fields": [
            {"key": "ticker",          "label": "Ticker",               "type": "text",   "required": True},
            {"key": "fecha_emision",   "label": "Fecha emisión",        "type": "date",   "required": True},
            {"key": "fecha_vencimiento","label": "Vencimiento",         "type": "date",   "required": True},
            {"key": "cupon anual %",   "label": "Cupón anual %",        "type": "text",
             "help": "Vacío para zero-coupon (mayoría de los DL)"},
            {"key": "frecuencia pagos","label": "Frecuencia pagos /año","type": "number", "step": "1",
             "help": "Solo si paga cupones intermedios"},
            {"key": "base calculo",    "label": "Base cálculo",         "type": "select",
             "options": _BASE_CALCULO_OPTIONS},
            {"key": "tipo amortizacion","label": "Tipo amortización",   "type": "select",
             "options": _TIPO_AMORT_OPTIONS},
            {"key": "amort inicio",    "label": "Amort inicio",         "type": "date",
             "help": "Solo amortizing"},
            {"key": "amort cantidad",  "label": "Cant. cuotas amort.",  "type": "number", "step": "1"},
            {"key": "tc_inicial",      "label": "TC inicial",           "type": "number", "step": "0.0001",
             "help": "Tipo de cambio de emisión (pesos/USD)"},
        ],
    },
    "TAMAR": {
        "label": "TAMAR (PURO / DUAL / DUAL_CER_TAMAR)",
        "fields": [
            {"key": "ticker",          "label": "Ticker",               "type": "text",   "required": True},
            {"key": "tipo",            "label": "Tipo",                 "type": "select", "required": True,
             "options": ["PURO", "DUAL", "DUAL_CER_TAMAR"]},
            {"key": "fecha_emision",   "label": "Fecha emisión",        "type": "date",   "required": True},
            {"key": "fecha_vencimiento","label": "Vencimiento",         "type": "date",   "required": True},
            {"key": "cupon anual %",   "label": "Cupón anual % (si aplica)", "type": "text",
             "help": "Solo para DUAL con cupón fijo adicional"},
            {"key": "frecuencia pagos","label": "Frecuencia pagos /año","type": "number", "step": "1"},
            {"key": "base calculo",    "label": "Base cálculo",         "type": "select",
             "options": _BASE_CALCULO_OPTIONS,
             "help": "TAMAR usa 30/360 por documento BONTE"},
            {"key": "tipo amortizacion","label": "Tipo amortización",   "type": "select",
             "options": _TIPO_AMORT_OPTIONS},
            {"key": "amort inicio",    "label": "Amort inicio",         "type": "date"},
            {"key": "amort cantidad",  "label": "Cant. cuotas amort.",  "type": "number", "step": "1"},
            {"key": "tasa_fija_mensual","label": "Tasa fija mensual (decimal)", "type": "number",
             "step": "0.0001", "help": "Solo DUAL — el floor mensual"},
            {"key": "spread",          "label": "Spread TAMAR (decimal)","type": "number",
             "step": "0.0001", "help": "0.05 = TAMAR + 5%"},
            {"key": "cer_base",        "label": "CER base",             "type": "number", "step": "0.01",
             "help": "Solo DUAL_CER_TAMAR — CER 10h pre-emisión"},
            {"key": "cer_spread",      "label": "Spread CER (decimal)", "type": "number",
             "step": "0.0001", "help": "Solo DUAL_CER_TAMAR"},
        ],
    },
}


def list_instruments(xlsx_path: str) -> List[Dict[str, str]]:
    """Return [{"ticker": "AL30D", "sheet": "Soberanos"}, ...]."""
    out = []
    with _LOCK:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            for sheet in _INSTRUMENT_SHEETS:
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                rows = ws.iter_rows(values_only=True)
                header = next(rows, None)
                if not header:
                    continue
                # Find ticker column (case-insensitive, stripped)
                norm = [str(c).lower().strip() if c else "" for c in header]
                if "ticker" not in norm:
                    continue
                tcol = norm.index("ticker")
                for row in rows:
                    if not row or row[tcol] is None:
                        continue
                    tk = str(row[tcol]).strip().upper()
                    if tk and tk != "NAN":
                        out.append({"ticker": tk, "sheet": sheet})
        finally:
            wb.close()
    return sorted(out, key=lambda x: x["ticker"])


def get_instrument(xlsx_path: str, ticker: str) -> Optional[Dict[str, Any]]:
    """Return {"sheet": str, "fields": {col: val}} for a ticker, None if not found."""
    ticker_u = ticker.upper().strip()
    with _LOCK:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        try:
            for sheet in _INSTRUMENT_SHEETS:
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
                norm = [h.lower() for h in hdr]
                if "ticker" not in norm:
                    continue
                tcol = norm.index("ticker") + 1
                for r in range(2, ws.max_row + 1):
                    tk = ws.cell(row=r, column=tcol).value
                    if tk and str(tk).strip().upper() == ticker_u:
                        fields = {}
                        for c, col_name in enumerate(hdr, start=1):
                            v = ws.cell(row=r, column=c).value
                            # JSON-friendly: dates → ISO strings, None passes through
                            if hasattr(v, "strftime"):
                                v = v.strftime("%Y-%m-%d")
                            fields[col_name.lower().strip()] = v
                        # Cashflows inline: el frontend renderiza una tabla
                        # editable al lado del form. Si no hay en hoja,
                        # devolvemos el synth (marcado para que el UI avise).
                        cf_payload = get_cashflows(xlsx_path, ticker_u, fallback_fields=fields)
                        return {
                            "sheet": sheet,
                            "fields": fields,
                            "cashflows": cf_payload["rows"],
                            "cashflows_source": cf_payload["source"],
                        }
        finally:
            wb.close()
    return None


def save_instrument(xlsx_path: str, sheet: str, fields: Dict[str, Any],
                    cashflows: Optional[List[Dict[str, Any]]] = None) -> Dict[str, str]:
    """Create or update a row by ticker in `sheet`. Returns {"action": "created"|"updated"}.

    Empty / None values are written as blank (cell cleared). Strings are stripped.
    Tickers are uppercased.

    TRANSACCIONAL: si `cashflows` se pasa, las dos escrituras (row del sheet
    + filas de Cashflows) ocurren en el MISMO workbook handle y se persisten
    con un único atomic save (.tmp + os.replace). Si la operación falla a
    mitad, ninguna mutación se persiste. Antes, las escrituras eran 2 saves
    separados y podía quedar el row sin cashflows (estado inconsistente).

    Pasar `cashflows=None` = no tocar la hoja Cashflows (preserva los
    cashflows existentes, o el synth-fallback si no hay).
    """
    if sheet not in SHEET_SCHEMAS:
        raise ValueError(f"Unknown sheet '{sheet}'. Allowed: {list(SHEET_SCHEMAS)}")
    ticker = str(fields.get("ticker", "")).strip().upper()
    if not ticker:
        raise ValueError("ticker is required")

    # Validar cashflows ANTES de abrir el workbook — fail fast sin tocar disco.
    parsed_cfs = _parse_cashflows(cashflows) if cashflows is not None else None

    with _LOCK:
        wb = openpyxl.load_workbook(xlsx_path)
        try:
            action = _write_instrument_row_on_wb(wb, sheet, ticker, fields)
            if parsed_cfs is not None:
                _write_cashflows_on_wb(wb, ticker, parsed_cfs)
            _atomic_save_workbook(wb, xlsx_path)
            logger.info(
                f"ABM: {action} {ticker} in {sheet}"
                + (f" · {len(parsed_cfs)} cashflows" if parsed_cfs is not None else "")
            )
        finally:
            wb.close()

    out: Dict[str, Any] = {"action": action, "ticker": ticker, "sheet": sheet}
    if parsed_cfs is not None:
        out["cashflows"] = len(parsed_cfs)
    return out


def _write_instrument_row_on_wb(wb, sheet: str, ticker: str, fields: Dict[str, Any]) -> str:
    """Mutar el workbook (en memoria) creando o actualizando la fila del ticker
    en `sheet`. Devuelve 'created' | 'updated'. No guarda — el caller decide."""
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found in workbook")
    ws = wb[sheet]
    hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
    norm = [h.lower() for h in hdr]
    if "ticker" not in norm:
        raise ValueError(f"Sheet '{sheet}' has no ticker column")
    tcol = norm.index("ticker") + 1

    row_idx = None
    for r in range(2, ws.max_row + 1):
        tk = ws.cell(row=r, column=tcol).value
        if tk and str(tk).strip().upper() == ticker:
            row_idx = r
            break
    action = "updated" if row_idx else "created"
    if row_idx is None:
        row_idx = ws.max_row + 1

    for c, col_name in enumerate(hdr, start=1):
        key = col_name.lower().strip()
        if key not in fields:
            continue
        val = fields[key]
        if isinstance(val, str):
            val = val.strip() or None
        if key == "ticker" and val is not None:
            val = str(val).upper()
        ws.cell(row=row_idx, column=c, value=val)
    return action


def delete_instrument(xlsx_path: str, ticker: str) -> Dict[str, str]:
    """Delete the row matching `ticker` from whatever sheet it lives in."""
    ticker_u = ticker.upper().strip()
    if not ticker_u:
        raise ValueError("ticker is required")

    with _LOCK:
        wb = openpyxl.load_workbook(xlsx_path)
        try:
            for sheet in _INSTRUMENT_SHEETS:
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                hdr = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                if "ticker" not in hdr:
                    continue
                tcol = hdr.index("ticker") + 1
                for r in range(2, ws.max_row + 1):
                    tk = ws.cell(row=r, column=tcol).value
                    if tk and str(tk).strip().upper() == ticker_u:
                        ws.delete_rows(r, 1)
                        # También limpiar cashflows huérfanos del ticker en
                        # la misma transacción atómica.
                        _write_cashflows_on_wb(wb, ticker_u, [])
                        _atomic_save_workbook(wb, xlsx_path)
                        logger.info(f"ABM: deleted {ticker_u} from {sheet}")
                        return {"action": "deleted", "ticker": ticker_u, "sheet": sheet}
        finally:
            wb.close()
    return {"action": "not_found", "ticker": ticker_u}


# Candidatos de nombre de columna para fecha de vencimiento en cada hoja.
_MATURITY_COL_CANDIDATES = frozenset([
    "fecha_vencimiento", "fecha vencimiento", "fecha_pago", "maturity",
])


def _parse_maturity(val) -> Optional[date]:
    """Intenta devolver un date a partir del valor de celda de openpyxl."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def purge_matured_instruments(xlsx_path: str) -> List[Dict[str, str]]:
    """Elimina del master Excel los instrumentos cuya fecha de vencimiento
    sea anterior a hoy.

    Se llama en startup, antes de que el repositorio cargue los datos, de
    modo que el repo ya arranca con la tabla limpia.

    Devuelve una lista de {"ticker", "sheet", "maturity"} por cada baja.
    """
    today = date.today()
    expired: List[tuple] = []

    with _LOCK:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        try:
            for sheet in _INSTRUMENT_SHEETS:
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not hdr:
                    continue
                norm = [str(c).lower().strip() if c is not None else "" for c in hdr]
                if "ticker" not in norm:
                    continue
                tcol = norm.index("ticker")

                mcol = next(
                    (norm.index(cand) for cand in _MATURITY_COL_CANDIDATES if cand in norm),
                    None,
                )
                if mcol is None:
                    continue

                for row in ws.iter_rows(min_row=2, values_only=True):
                    tk = row[tcol]
                    if not tk:
                        continue
                    tk_str = str(tk).strip().upper()
                    if not tk_str or tk_str == "NAN":
                        continue
                    m_date = _parse_maturity(row[mcol])
                    if m_date and m_date < today:
                        expired.append((tk_str, sheet, m_date))
        finally:
            wb.close()

    deleted = []
    for ticker, sheet, m_date in expired:
        result = delete_instrument(xlsx_path, ticker)
        if result.get("action") == "deleted":
            deleted.append({
                "ticker": ticker,
                "sheet": sheet,
                "maturity": m_date.isoformat(),
            })
            logger.info(
                "Startup purge: eliminado %s (%s, vto %s)", ticker, sheet, m_date
            )

    return deleted
