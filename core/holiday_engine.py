"""
holiday_engine.py — Calendario BYMA y feriados Argentina unificado
===================================================================
Integra pandas_market_calendars (XBUE) con 5 fuentes de feriados
y funciones de calendario para renta fija argentina.

HORARIO BYMA:  10:30 - 17:00 (America/Argentina/Buenos_Aires)
SETTLEMENT:    T+0 (CI) | T+1 (bonos, cauciones, LECAP)
               T+2 eliminado — ya no se usa en BYMA.

FUENTES DE FERIADOS (orden de prioridad):
  1. nolaborables.com.ar/api/v2/feriados/{año}   — Boletin Oficial AR
  2. api.argentinadatos.com/v1/feriados/{año}     — datos.gob.ar
  3. date.nager.at/api/v3/PublicHolidays/{año}/AR — internacional
  4. holidays (Python pura, sin red)              — fallback offline
  5. XBUE exchange_calendars                      — base pandas_market_calendars

FUNCIONES CALENDARIO:
  is_habil()             — es dia habil BYMA?
  date_range_habil()     — rango de dias habiles
  siguiente_habil()      — proximo dia habil
  anterior_habil()       — dia habil anterior
  settlement_byma()      — fecha liquidacion T+0 / T+1
  fechas_cupon()         — schedule de cupones ajustado por BYMA
  bdays_hasta()          — dias habiles faltantes
  is_open_now()          — estado del mercado (10:30-17:00)
  accrual_dias()         — dias corridos para accrual de interes

FERIADOS:
  fetch / merge / refresh / cache / Excel con 4 hojas

pip install pandas pandas_market_calendars holidays requests openpyxl scipy
"""

from __future__ import annotations

import json
import logging
from datetime import date, datetime, timedelta, time
from functools import lru_cache
from pathlib import Path

import holidays as holidays_lib
import pandas as pd
import httpx

logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")
log = logging.getLogger("holiday_engine")


# ═════════════════════════════════════════════════════════════════
#  CONSTANTES
# ═════════════════════════════════════════════════════════════════
TZ_BA           = "America/Argentina/Buenos_Aires"
BYMA_OPEN       = time(10, 30)
BYMA_CLOSE      = time(17, 0)

DATA_DIR = Path(__file__).parent.parent / "data"
EXCEL_PATH      = DATA_DIR / "feriados_ar.xlsx"
CACHE_JSON      = DATA_DIR / "feriados_ar_cache.json"
REFRESH_DAYS    = 60
TIMEOUT_API     = 10
AÑOS_COBERTURA  = range(2020, 2030)

# Colores Excel
C_HEADER    = "042C53"
C_WHITE     = "FFFFFF"
C_GREEN_BG  = "EAF3DE"
C_AMBER_BG  = "FAEEDA"
C_PURPLE_BG = "EEEDFE"
C_BLUE_BG   = "E6F1FB"
C_GRAY_BG   = "F1EFE8"
C_GREEN_FG  = "27500A"
C_AMBER_FG  = "633806"
C_PURPLE_FG = "3C3489"
C_BLUE_FG   = "0C447C"
C_DARK_FG   = "2C2C2A"

TIPO_STYLE = {
    "inamovible":  (C_GREEN_BG,  C_GREEN_FG),
    "puente":      (C_AMBER_BG,  C_AMBER_FG),
    "trasladable": (C_PURPLE_BG, C_PURPLE_FG),
    "nolaborable": (C_BLUE_BG,   C_BLUE_FG),
    "bancario":    (C_BLUE_BG,   C_BLUE_FG),
}

PRIORIDAD_FUENTE = [
    "nolaborables.com.ar",
    "argentinadatos.com",
    "nager.date",
    "holidays_lib",
    "xbue_pmc",
]


# ═════════════════════════════════════════════════════════════════
#  CALENDARIO BYMA — instancia cacheada
# ═════════════════════════════════════════════════════════════════
@lru_cache(maxsize=1)
def _get_byma() -> "mcal.MarketCalendar":
    import pandas_market_calendars as mcal
    return mcal.get_calendar("XBUE")


# ═════════════════════════════════════════════════════════════════
#  FERIADOS AR — fuente de verdad para días hábiles
# ═════════════════════════════════════════════════════════════════
# Prioridad (decisión del usuario): el Excel oficial de feriados (armado desde
# las APIs nolaborables/argentinadatos) manda; el paquete pandas (XBUE) queda
# como fallback para años sin cobertura del Excel.
#
# El Excel es un MERGE de 5 fuentes y arrastra fechas espurias cuya única fuente
# es `xbue_pmc` (pandas_market_calendars ubica mal trasladables/puentes, p.ej.
# 2025-11-17). Se descartan esas filas: confiamos en API + holidays_lib/nager,
# nunca en XBUE para AGREGAR feriados. Ver memoria xbue-calendar-missing-ar-holidays.

@lru_cache(maxsize=1)
def _ar_holidays() -> frozenset:
    """Feriados AR (no laborables BYMA) desde data/feriados_ar.xlsx, excluyendo
    las filas cuya fuente principal es solo `xbue_pmc`. frozenset[date]."""
    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name="Feriados AR")
        fechas = pd.to_datetime(df["Fecha"], errors="coerce")
        fuente = df["Fuente principal"].astype(str).str.strip()
        days = {
            ts.date()
            for ts, src in zip(fechas, fuente)
            if pd.notna(ts) and src != "xbue_pmc"
        }
        return frozenset(days)
    except Exception as e:  # Excel ausente/corrupto → todo al fallback XBUE
        log.warning(f"No se pudo leer feriados AR de {EXCEL_PATH}: {e}")
        return frozenset()


@lru_cache(maxsize=1)
def _ar_cobertura() -> tuple[int, int]:
    """Rango de años (min, max) cubierto por el Excel de feriados. (0, -1) si vacío."""
    ys = {d.year for d in _ar_holidays()}
    return (min(ys), max(ys)) if ys else (0, -1)


def _xbue_habil(d: date) -> bool:
    """Fallback: ¿día de rueda según el calendario XBUE de pandas?"""
    s = _get_byma().schedule(d.strftime("%Y-%m-%d"), d.strftime("%Y-%m-%d"))
    return not s.empty


def _es_habil(d: date) -> bool:
    """Día hábil BYMA. Fin de semana → no. Para años cubiertos por el Excel
    oficial AR, hábil = no es feriado de esa lista; fuera de cobertura, fallback
    al calendario XBUE de pandas."""
    if d.weekday() >= 5:  # 5=sábado, 6=domingo
        return False
    lo, hi = _ar_cobertura()
    if lo <= d.year <= hi:
        return d not in _ar_holidays()
    return _xbue_habil(d)


# ═════════════════════════════════════════════════════════════════
#  FUNCIONES DE CALENDARIO
# ═════════════════════════════════════════════════════════════════

def es_habil(d: date) -> bool:
    """Día hábil BYMA — API **date-native** (O(1): weekday + lookup en el frozenset
    de feriados). PREFERIR sobre `is_habil(str)` en hot-paths (cer_reference_date,
    settlement): evita el round-trip `date→str→pd.Timestamp→date` que cuesta ~2ms
    por la creación del Timestamp; date-native es ~1µs."""
    return _es_habil(d)


def is_habil(fecha: str) -> bool:
    """True si la fecha (str) es dia habil BYMA. Wrapper de compat sobre `es_habil`."""
    return es_habil(pd.Timestamp(fecha).date())


def date_range_habil(desde: str, hasta: str) -> pd.DatetimeIndex:
    """DatetimeIndex con todos los dias habiles BYMA en el rango."""
    dias = pd.date_range(start=desde, end=hasta, freq="D")
    return pd.DatetimeIndex([d for d in dias if _es_habil(d.date())])


def _buscar_habil(fecha: str, direccion: int, inclusive: bool) -> pd.Timestamp:
    """Busca el dia habil mas cercano en la direccion dada (+1 o -1)."""
    ts = pd.Timestamp(fecha)
    start = 0 if inclusive else 1
    for i in range(start, 15):
        cand = ts + pd.Timedelta(days=i * direccion)
        if _es_habil(cand.date()):
            return cand
    raise ValueError(f"No se encontro dia habil cerca de {fecha}")


def siguiente_habil(fecha: str, inclusive: bool = True) -> pd.Timestamp:
    """Siguiente (o mismo si inclusive) dia habil BYMA."""
    return _buscar_habil(fecha, +1, inclusive)


def anterior_habil(fecha: str, inclusive: bool = True) -> pd.Timestamp:
    """Anterior (o mismo si inclusive) dia habil BYMA."""
    return _buscar_habil(fecha, -1, inclusive)


def settlement_byma_date(trade: date, lag: int = 1) -> date:
    """Liquidación BYMA T+N **date-native** (sin parseo de string). Ver
    `settlement_byma` para la convención. Preferir en hot-paths."""
    if lag not in (0, 1):
        raise ValueError(f"BYMA solo soporta T+0 y T+1. Recibido lag={lag}")
    cand = trade
    if lag == 0:  # CI: mismo día si hábil, si no el siguiente hábil
        for _ in range(15):
            if _es_habil(cand):
                return cand
            cand += timedelta(days=1)
        raise ValueError(f"No se encontró día hábil cerca de {trade}")
    count = 0
    while count < lag:
        cand += timedelta(days=1)
        if _es_habil(cand):
            count += 1
    return cand


def settlement_byma(trade_date: str, lag: int = 1) -> pd.Timestamp:
    """
    Fecha de liquidacion BYMA T+N contando solo dias habiles.

    BYMA liquida:
      T+0  contado inmediato (CI)
      T+1  bonos, LECAP, cauciones, ONs  (default)

    >>> settlement_byma("2026-04-10", lag=1)  # Viernes -> lunes
    Timestamp('2026-04-14 00:00:00')
    >>> settlement_byma("2026-04-10", lag=0)  # CI mismo dia
    Timestamp('2026-04-10 00:00:00')

    Wrapper str→pd.Timestamp sobre `settlement_byma_date` (date-native).
    """
    return pd.Timestamp(settlement_byma_date(pd.Timestamp(trade_date).date(), lag))


def fechas_cupon(
    emision: str,
    vencimiento: str,
    freq_meses: int = 6,
    convencion: str = "following",
) -> pd.DataFrame:
    """
    Schedule de cupones ajustado por calendario BYMA.

    Convenciones:
      "following"          — proximo habil si cae en feriado
      "modified_following" — habil anterior si cambiaria de mes
    """
    fechas_nominales = pd.date_range(
        start=pd.Timestamp(emision) + pd.DateOffset(months=freq_meses),
        end=pd.Timestamp(vencimiento),
        freq=pd.DateOffset(months=freq_meses),
    )

    rows = []
    for i, fn in enumerate(fechas_nominales, 1):
        fn_str = fn.strftime("%Y-%m-%d")
        if not is_habil(fn_str):
            if convencion == "modified_following":
                fa = siguiente_habil(fn_str)
                if fa.month != fn.month:
                    fa = anterior_habil(fn_str)
            else:
                fa = siguiente_habil(fn_str)
        else:
            fa = fn

        rows.append({
            "n_cupon":        i,
            "fecha_nominal":  fn.strftime("%d/%m/%Y"),
            "fecha_ajustada": fa.strftime("%d/%m/%Y"),
            "dia_semana":     fa.strftime("%A"),
            "ajustado":       "SI" if fa != fn else "-",
        })

    return pd.DataFrame(rows)


def bdays_hasta(desde: str, hasta: str) -> int:
    """Dias habiles BYMA entre dos fechas (inclusive)."""
    return len(date_range_habil(desde, hasta))


def is_open_now() -> dict:
    """
    Estado actual del mercado BYMA.
    Horario: 10:30 - 17:00 Buenos Aires.
    """
    now = pd.Timestamp.now(tz=TZ_BA)
    hoy = now.strftime("%Y-%m-%d")
    es_habil = is_habil(hoy)

    hora_actual = now.time()
    abierto = es_habil and BYMA_OPEN <= hora_actual <= BYMA_CLOSE

    return {
        "abierto":      abierto,
        "hora_ba":      now.strftime("%H:%M"),
        "apertura":     BYMA_OPEN.strftime("%H:%M"),
        "cierre":       BYMA_CLOSE.strftime("%H:%M"),
        "es_habil_hoy": es_habil,
    }


def accrual_dias(
    ultimo_cupon: str,
    fecha_calculo: str,
    convencion: str = "30/360",
) -> dict:
    """
    Dias corridos para accrual de interes.

    Convenciones: "30/360", "Act/365", "Act/360", "Act/habil"
    """
    t0 = pd.Timestamp(ultimo_cupon)
    t1 = pd.Timestamp(fecha_calculo)

    if convencion == "30/360":
        d0, m0, y0 = t0.day, t0.month, t0.year
        d1, m1, y1 = t1.day, t1.month, t1.year
        d0 = min(d0, 30)
        if d0 == 30:
            d1 = min(d1, 30)
        dias = 360 * (y1 - y0) + 30 * (m1 - m0) + (d1 - d0)
        base = 360
    elif convencion == "Act/365":
        dias = (t1 - t0).days
        base = 365
    elif convencion == "Act/360":
        dias = (t1 - t0).days
        base = 360
    elif convencion == "Act/habil":
        dias = bdays_hasta(ultimo_cupon, fecha_calculo)
        base = 252
    else:
        raise ValueError(f"Convencion desconocida: {convencion}")

    return {
        "convencion":       convencion,
        "dias":             dias,
        "base":             base,
        "fraccion_periodo": round(dias / base * 2, 8),
        "accrued_factor":   round(dias / base, 8),
    }


# ═════════════════════════════════════════════════════════════════
#  FUENTES DE FERIADOS
# ═════════════════════════════════════════════════════════════════

def fetch_nolaborables(año: int) -> list[dict]:
    """Fuente 1: nolaborables.com.ar — Boletin Oficial AR."""
    url = f"https://nolaborables.com.ar/api/v2/feriados/{año}"
    try:
        resp = httpx.get(url, timeout=TIMEOUT_API, follow_redirects=True)
        resp.raise_for_status()
        data = resp.json()
        result = []
        for h in data:
            try:
                fecha = pd.Timestamp(f"{año}-{int(h['mes']):02d}-{int(h['dia']):02d}")
                result.append({
                    "fecha":  fecha,
                    "nombre": h.get("motivo", ""),
                    "tipo":   h.get("tipo", ""),
                    "año":    año,
                    "fuente": "nolaborables.com.ar",
                })
            except Exception:
                pass
        log.info(f"nolaborables.com.ar  {año}: {len(result):3d} feriados")
        return result
    except Exception as e:
        log.warning(f"nolaborables.com.ar  {año}: NO DISPONIBLE ({e})")
        return []


def fetch_argentinadatos(año: int) -> list[dict]:
    """Fuente 2: api.argentinadatos.com — datos.gob.ar."""
    url = f"https://api.argentinadatos.com/v1/feriados/{año}"
    try:
        resp = httpx.get(url, timeout=TIMEOUT_API, follow_redirects=True)
        resp.raise_for_status()
        data = resp.json()
        result = []
        for h in data:
            try:
                result.append({
                    "fecha":  pd.Timestamp(h["fecha"]),
                    "nombre": h.get("nombre", ""),
                    "tipo":   h.get("tipo", ""),
                    "año":    año,
                    "fuente": "argentinadatos.com",
                })
            except Exception:
                pass
        log.info(f"argentinadatos.com   {año}: {len(result):3d} feriados")
        return result
    except Exception as e:
        log.warning(f"argentinadatos.com   {año}: NO DISPONIBLE ({e})")
        return []


def fetch_nager(año: int) -> list[dict]:
    """Fuente 3: date.nager.at — internacional."""
    url = f"https://date.nager.at/api/v3/PublicHolidays/{año}/AR"
    try:
        resp = httpx.get(url, timeout=TIMEOUT_API, follow_redirects=True)
        resp.raise_for_status()
        data = resp.json()
        result = []
        for h in data:
            try:
                result.append({
                    "fecha":  pd.Timestamp(h["date"]),
                    "nombre": h.get("localName", h.get("name", "")),
                    "tipo":   "inamovible",
                    "año":    año,
                    "fuente": "nager.date",
                })
            except Exception:
                pass
        log.info(f"nager.date           {año}: {len(result):3d} feriados")
        return result
    except Exception as e:
        log.warning(f"nager.date           {año}: NO DISPONIBLE ({e})")
        return []


def fetch_holidays_lib(año: int) -> list[dict]:
    """Fuente 4: holidays lib — Python pura, sin red."""
    ar = holidays_lib.Argentina(years=año)
    result = []
    for fecha, nombre in ar.items():
        tipo = "puente" if "turístico" in nombre.lower() else "inamovible"
        result.append({
            "fecha":  pd.Timestamp(fecha),
            "nombre": nombre,
            "tipo":   tipo,
            "año":    año,
            "fuente": "holidays_lib",
        })
    log.info(f"holidays_lib         {año}: {len(result):3d} feriados")
    return result


def fetch_xbue(año: int) -> list[dict]:
    """Fuente 5: XBUE exchange_calendars — base pandas_market_calendars."""
    cal = _get_byma()
    hols = cal.holidays().holidays
    result = []
    for h in hols:
        ts = pd.Timestamp(h)
        if ts.year == año:
            result.append({
                "fecha":  ts,
                "nombre": "",
                "tipo":   "inamovible",
                "año":    año,
                "fuente": "xbue_pmc",
            })
    log.info(f"xbue_pmc             {año}: {len(result):3d} feriados")
    return result


# ═════════════════════════════════════════════════════════════════
#  MERGE — combina todas las fuentes
# ═════════════════════════════════════════════════════════════════

def merge_fuentes(listas: list[list[dict]]) -> pd.DataFrame:
    """
    Combina registros de multiples fuentes.
    Una misma fecha puede estar en varias — se consolida con la fuente
    mas confiable y todas las fuentes como lista.
    """
    por_fecha: dict[pd.Timestamp, dict] = {}

    for lista in listas:
        for h in lista:
            fecha = h["fecha"]
            if fecha not in por_fecha:
                por_fecha[fecha] = h.copy()
                por_fecha[fecha]["fuentes"] = [h["fuente"]]
            else:
                existing = por_fecha[fecha]
                if h["fuente"] not in existing["fuentes"]:
                    existing["fuentes"].append(h["fuente"])
                pri_new = PRIORIDAD_FUENTE.index(h["fuente"]) if h["fuente"] in PRIORIDAD_FUENTE else 99
                pri_ex  = PRIORIDAD_FUENTE.index(existing["fuente"]) if existing["fuente"] in PRIORIDAD_FUENTE else 99
                if pri_new < pri_ex:
                    existing["nombre"] = h["nombre"]
                    existing["tipo"]   = h["tipo"]
                    existing["fuente"] = h["fuente"]

    rows = []
    for fecha, h in sorted(por_fecha.items()):
        fuentes_str = " + ".join(h["fuentes"])
        confirmado  = len(h["fuentes"]) >= 2
        rows.append({
            "fecha":         fecha,
            "dia_semana":    fecha.strftime("%A"),
            "nombre":        h["nombre"],
            "tipo":          h["tipo"],
            "año":           int(h["año"]),
            "fuente_ppal":   h["fuente"],
            "todas_fuentes": fuentes_str,
            "confirmado":    "Si" if confirmado else "Solo en 1 fuente",
            "n_fuentes":     len(h["fuentes"]),
        })

    return pd.DataFrame(rows)


# ═════════════════════════════════════════════════════════════════
#  DESCARGA COMPLETA
# ═════════════════════════════════════════════════════════════════

def descargar_todos(años=AÑOS_COBERTURA) -> pd.DataFrame:
    """Descarga feriados de todas las fuentes para todos los anos."""
    log.info(f"Iniciando descarga para años {list(años)[0]}-{list(años)[-1]}")
    todas = []

    for año in años:
        log.info(f"── Año {año} ──")
        listas_año = []

        r1 = fetch_nolaborables(año)
        if r1:
            listas_año.append(r1)

        r2 = fetch_argentinadatos(año)
        if r2:
            listas_año.append(r2)

        r3 = fetch_nager(año)
        if r3:
            listas_año.append(r3)

        listas_año.append(fetch_holidays_lib(año))
        listas_año.append(fetch_xbue(año))

        todas.extend(listas_año)

    df = merge_fuentes(todas)
    log.info(f"Total feriados unicos: {len(df)}")
    return df


# ═════════════════════════════════════════════════════════════════
#  EXCEL — generacion con formato
# ═════════════════════════════════════════════════════════════════

def _thin_border():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="B4B2A9")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row, col, value, width=None):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", color=C_WHITE, bold=True, size=10)
    c.fill      = PatternFill("solid", fgColor=C_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    c.border    = _thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def generar_excel(df: pd.DataFrame, path: Path = EXCEL_PATH) -> Path:
    """
    Genera feriados_ar.xlsx con 4 hojas:
      1. Feriados AR    — tabla principal
      2. Metadata       — fecha de update y configuracion
      3. Leyenda tipos  — definicion de cada tipo BYMA
      4. Por año        — pivot resumen
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()

    # ── Hoja 1: Feriados AR ──
    ws1 = wb.active
    ws1.title = "Feriados AR"

    headers = ["#", "Fecha", "Dia", "Nombre del feriado", "Tipo", "Año",
               "Fuente principal", "Todas las fuentes", "Confirmado", "N fuentes"]
    widths  = [5,   14,      12,    42,                   14,      7,
               22,              36,                   13,          11]

    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        _header_cell(ws1, 1, ci, h, w)
    ws1.row_dimensions[1].height = 22

    for i, row in df.iterrows():
        ri       = i + 2
        tipo     = row.get("tipo", "")
        bg, fg   = TIPO_STYLE.get(tipo, (C_GRAY_BG if i % 2 == 0 else C_WHITE, C_DARK_FG))
        fill_row = PatternFill("solid", fgColor=bg) if bg else None
        font_row = Font(name="Arial", color=fg, size=10)
        border   = _thin_border()
        center   = Alignment(horizontal="center", vertical="center")
        left     = Alignment(horizontal="left",   vertical="center")

        vals = [
            i + 1,
            row["fecha"].to_pydatetime().date(),
            row["dia_semana"],
            row["nombre"],
            row["tipo"],
            int(row["año"]),
            row["fuente_ppal"],
            row["todas_fuentes"],
            row["confirmado"],
            int(row["n_fuentes"]),
        ]
        aligns = [center, center, center, left, center, center, left, left, center, center]

        for ci, (val, aln) in enumerate(zip(vals, aligns), 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.font      = font_row
            cell.alignment = aln
            cell.border    = border
            if fill_row:
                cell.fill = fill_row
            if ci == 2 and hasattr(val, "strftime"):
                cell.number_format = "DD/MM/YYYY"

        ws1.row_dimensions[ri].height = 17

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df)+1}"

    # ── Hoja 2: Metadata ──
    ws2 = wb.create_sheet("Metadata")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 52

    ahora        = datetime.now()
    prox_refresh = (ahora + timedelta(days=REFRESH_DAYS)).strftime("%Y-%m-%d")
    años_str     = f"{df['año'].min()} - {df['año'].max()}"

    meta_rows = [
        ("Archivo",                     str(path.name)),
        ("Ultima actualizacion",        ahora.strftime("%Y-%m-%d %H:%M:%S")),
        ("Proximo refresh recomendado", prox_refresh),
        ("Frecuencia de refresh",       f"Cada {REFRESH_DAYS} dias"),
        ("Total feriados",              f"{len(df)}"),
        ("Anos cubiertos",              años_str),
        ("Horario BYMA",               f"{BYMA_OPEN.strftime('%H:%M')} - {BYMA_CLOSE.strftime('%H:%M')}"),
        ("Settlement default",          "T+1 (T+2 eliminado)"),
        ("Fuente 1",                    "nolaborables.com.ar/api/v2/feriados/{año}"),
        ("Fuente 2",                    "api.argentinadatos.com/v1/feriados/{año}"),
        ("Fuente 3",                    "date.nager.at/api/v3/PublicHolidays/{año}/AR"),
        ("Fuente 4 (offline)",          "holidays lib (Python)"),
        ("Fuente 5 (offline)",          "XBUE exchange_calendars"),
        ("Script de refresh",           "python holiday_engine.py --refresh"),
    ]

    _header_cell(ws2, 1, 1, "Parametro", 30)
    _header_cell(ws2, 1, 2, "Valor",     52)
    ws2.row_dimensions[1].height = 20

    for ri, (k, v) in enumerate(meta_rows, 2):
        ck = ws2.cell(row=ri, column=1, value=k)
        cv = ws2.cell(row=ri, column=2, value=str(v))
        ck.font = Font(name="Arial", bold=True,  size=10, color=C_DARK_FG)
        cv.font = Font(name="Arial", bold=False, size=10, color=C_DARK_FG)
        if k == "Ultima actualizacion":
            cv.font = Font(name="Arial", bold=True, size=10, color=C_AMBER_FG)
        if k == "Proximo refresh recomendado":
            cv.font = Font(name="Arial", bold=True, size=10, color="A32D2D")
        ck.border = cv.border = _thin_border()
        if ri % 2 == 0:
            ck.fill = cv.fill = PatternFill("solid", fgColor=C_GRAY_BG)
        ws2.row_dimensions[ri].height = 17

    # ── Hoja 3: Leyenda tipos ──
    ws3 = wb.create_sheet("Leyenda tipos")
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["C"].width = 20

    _header_cell(ws3, 1, 1, "Tipo",        16)
    _header_cell(ws3, 1, 2, "Descripcion", 50)
    _header_cell(ws3, 1, 3, "Ejemplo",     20)
    ws3.row_dimensions[1].height = 20

    leyenda = [
        ("inamovible",  "Fecha fija por ley — nunca se traslada",           "Año Nuevo, Navidad"),
        ("puente",      "Decretado por el PE — puede variar cada año",      "Puentes turisticos"),
        ("trasladable", "Se mueve al lunes mas cercano si cae mar-mie",     "San Martin, Belgrano"),
        ("nolaborable", "Optativo — el empleador decide si se trabaja",     "Semana Financiera"),
        ("bancario",    "Solo afecta a entidades financieras (BCRA)",       "Sin equivalente"),
    ]
    for ri, (tipo, desc, ejemplo) in enumerate(leyenda, 2):
        bg, fg = TIPO_STYLE.get(tipo, (C_GRAY_BG, C_DARK_FG))
        fill   = PatternFill("solid", fgColor=bg)
        font   = Font(name="Arial", color=fg, size=10)
        border = _thin_border()
        for ci, val in enumerate([tipo, desc, ejemplo], 1):
            c = ws3.cell(row=ri, column=ci, value=val)
            c.fill      = fill
            c.font      = font
            c.border    = border
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws3.row_dimensions[ri].height = 17

    # ── Hoja 4: Por año ──
    ws4 = wb.create_sheet("Por año")
    for col_letter in "ABCDEF":
        ws4.column_dimensions[col_letter].width = 14

    pivot = (df.groupby(["año", "tipo"])
               .size()
               .unstack(fill_value=0)
               .reset_index())
    pivot["total"] = pivot.drop("año", axis=1).sum(axis=1)
    cols_pivot = ["año"] + [c for c in pivot.columns if c not in ["año", "total"]] + ["total"]
    pivot = pivot[cols_pivot]

    for ci, col in enumerate(cols_pivot, 1):
        _header_cell(ws4, 1, ci, str(col).capitalize(),
                     ws4.column_dimensions[get_column_letter(ci)].width)

    for ri, (_, row) in enumerate(pivot.iterrows(), 2):
        for ci, col in enumerate(cols_pivot, 1):
            c = ws4.cell(row=ri, column=ci, value=int(row[col]))
            c.font      = Font(name="Arial", size=10, color=C_DARK_FG)
            c.alignment = Alignment(horizontal="center")
            c.border    = _thin_border()
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor=C_GRAY_BG)

    wb.save(path)
    log.info(f"Excel guardado: {path} ({path.stat().st_size // 1024} KB)")
    return path


# ═════════════════════════════════════════════════════════════════
#  CACHE JSON
# ═════════════════════════════════════════════════════════════════

def guardar_cache(df: pd.DataFrame, path: Path = CACHE_JSON):
    """Guarda DataFrame + timestamp en JSON para control de refresh."""
    data = {
        "ultima_actualizacion": datetime.now().isoformat(),
        "total_registros":      len(df),
        "años_cubiertos":       f"{df['año'].min()}-{df['año'].max()}",
        "feriados": [
            {
                "fecha":         row["fecha"].strftime("%Y-%m-%d"),
                "nombre":        row["nombre"],
                "tipo":          row["tipo"],
                "año":           int(row["año"]),
                "fuente_ppal":   row["fuente_ppal"],
                "todas_fuentes": row["todas_fuentes"],
                "confirmado":    row["confirmado"],
            }
            for _, row in df.iterrows()
        ],
    }
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info(f"Cache JSON guardado: {path}")


def cargar_cache(path: Path = CACHE_JSON) -> tuple[pd.DataFrame | None, datetime | None]:
    """Carga cache JSON. Retorna (DataFrame, fecha_actualizacion) o (None, None)."""
    if not path.exists():
        return None, None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        fecha_cache = datetime.fromisoformat(data["ultima_actualizacion"])
        df = pd.DataFrame(data["feriados"])
        df["fecha"] = pd.to_datetime(df["fecha"])
        log.info(f"Cache cargado: {len(df)} feriados, actualizado {fecha_cache.date()}")
        return df, fecha_cache
    except Exception as e:
        log.warning(f"Cache invalido: {e}")
        return None, None


# ═════════════════════════════════════════════════════════════════
#  REFRESH AUTOMATICO — logica de 60 dias
# ═════════════════════════════════════════════════════════════════

def necesita_refresh(path_cache: Path = CACHE_JSON,
                     path_excel: Path = EXCEL_PATH,
                     dias: int = REFRESH_DAYS) -> tuple[bool, str]:
    """Determina si hay que refrescar los datos."""
    _, fecha_cache = cargar_cache(path_cache)

    if fecha_cache is None and not path_excel.exists():
        return True, "No existe cache ni Excel — primera ejecucion"

    if fecha_cache is None:
        return True, "Cache no encontrado — re-descargar"

    dias_desde = (datetime.now() - fecha_cache).days

    if dias_desde >= dias:
        return True, f"Han pasado {dias_desde} dias (limite: {dias})"

    hoy = datetime.now()
    if hoy.month == 1 and hoy.day <= 3:
        return True, "Inicio de año — posibles feriados puente nuevos"

    return False, f"Cache vigente — actualizado hace {dias_desde} dias (proximo en {dias - dias_desde})"


def refresh(forzar: bool = False,
            años: range = AÑOS_COBERTURA,
            excel_path: Path = EXCEL_PATH,
            cache_path: Path = CACHE_JSON) -> pd.DataFrame:
    """
    Ciclo completo: verificar -> descargar -> excel -> cache.
    Retorna el DataFrame actualizado.
    """
    requiere, motivo = necesita_refresh(cache_path, excel_path)

    if not requiere and not forzar:
        log.info(f"Sin refresh necesario: {motivo}")
        df, _ = cargar_cache(cache_path)
        return df if df is not None else descargar_todos(años)

    log.info(f"Iniciando refresh: {motivo}")
    df = descargar_todos(años)

    log.info("Generando Excel...")
    generar_excel(df, excel_path)

    log.info("Guardando cache JSON...")
    guardar_cache(df, cache_path)

    log.info(f"Refresh completado: {len(df)} feriados, {excel_path}")
    return df


# ═════════════════════════════════════════════════════════════════
#  CONTROLES — python holiday_engine.py
# ═════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Holiday Engine — BYMA calendar + feriados AR")
    parser.add_argument("--refresh", action="store_true",
                        help="Revisar y refrescar si corresponde")
    parser.add_argument("--force",   action="store_true",
                        help="Forzar refresh ignorando los 60 dias")
    parser.add_argument("--status",  action="store_true",
                        help="Mostrar estado del cache")
    parser.add_argument("--control", action="store_true",
                        help="Ejecutar controles de calendario")
    args = parser.parse_args()

    if args.status:
        requiere, motivo = necesita_refresh()
        print(f"Estado: {'REFRESH REQUERIDO' if requiere else 'OK'}")
        print(f"Motivo: {motivo}")
        _, fecha = cargar_cache()
        if fecha:
            print(f"Ultimo update: {fecha.strftime('%Y-%m-%d %H:%M')}")
            print(f"Proximo refresh: {(fecha + timedelta(days=REFRESH_DAYS)).strftime('%Y-%m-%d')}")

    elif args.control:
        # ── CONTROLES DE CALENDARIO ──
        print("=" * 60)
        print("CONTROLES — holiday_engine.py")
        print("=" * 60)

        print(f"\n  Horario BYMA: {BYMA_OPEN.strftime('%H:%M')} - {BYMA_CLOSE.strftime('%H:%M')}")
        print("  Settlement default: T+1 (T+2 eliminado)")

        # Control 1: Semana Santa 2026
        print("\n[1] Semana Santa 2026:")
        for f in ["2026-04-01", "2026-04-02", "2026-04-03", "2026-04-06", "2026-04-07"]:
            ts = pd.Timestamp(f)
            print(f"    {f} {ts.strftime('%A'):10s} -> {'HABIL' if is_habil(f) else 'FERIADO'}")

        # Control 2: Settlement T+1
        print("\n[2] Settlement T+1 BYMA:")
        for d in ["2026-04-08", "2026-04-09", "2026-04-10"]:
            s1 = settlement_byma(d, lag=1)
            s0 = settlement_byma(d, lag=0)
            print(f"    Trade {d} -> T+0: {s0.strftime('%d/%m/%Y %A')}  |  T+1: {s1.strftime('%d/%m/%Y %A')}")

        # Control 3: Schedule cupones TX26
        print("\n[3] Schedule cupones TX26 (semestral):")
        df_cup = fechas_cupon("2021-11-09", "2027-11-09", freq_meses=6)
        print(df_cup.to_string(index=False))

        # Control 4: Time-to-maturity
        print("\n[4] Dias hasta vencimiento (desde 10/04/2026):")
        hoy = "2026-04-10"
        for bono, vto in [("TX26", "2026-11-09"), ("TX28", "2028-11-09"), ("TX31", "2031-03-23")]:
            dc = (pd.Timestamp(vto) - pd.Timestamp(hoy)).days
            dh = bdays_hasta(hoy, vto)
            print(f"    {bono}: {dc:4d} corridos | {dh:4d} habiles | ratio {dh/dc:.3f}")

        # Control 5: Estado mercado
        print("\n[5] Estado BYMA ahora:")
        for k, v in is_open_now().items():
            print(f"    {k:15s}: {v}")

        # Control 6: Accrual
        print("\n[6] Accrual TX26 — ult cupon 09/11/2025, hoy 10/04/2026:")
        acc = accrual_dias("2025-11-09", "2026-04-10", convencion="30/360")
        for k, v in acc.items():
            print(f"    {k:20s}: {v}")

        # Control 7: Feriados XBUE 2026
        print("\n[7] Feriados XBUE 2026:")
        cal = _get_byma()
        hols = sorted(set(h for h in cal.holidays().holidays if str(h)[:4] == "2026"))
        for h in hols:
            ts = pd.Timestamp(h)
            print(f"    {ts.strftime('%d/%m/%Y')} ({ts.strftime('%A')})")

    elif args.refresh or args.force:
        df = refresh(forzar=args.force)
        print(f"\nResultado: {len(df)} feriados | Excel: {EXCEL_PATH} | Cache: {CACHE_JSON}")

    else:
        # Default: refresh completo + resumen
        df = refresh(forzar=True)

        print("\n" + "=" * 60)
        print("Resumen por año:")
        print("=" * 60)
        resumen = df.groupby(["año", "tipo"]).size().unstack(fill_value=0)
        print(resumen.to_string())

        print(f"\nConfirmados (2+ fuentes): {(df['n_fuentes'] >= 2).sum()}")
        print(f"Solo 1 fuente (verificar): {(df['n_fuentes'] == 1).sum()}")
        print("\nArchivos generados:")
        print(f"  {EXCEL_PATH.absolute()}")
        print(f"  {CACHE_JSON.absolute()}")
