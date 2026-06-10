"""Exporta el panel de Obligaciones Negociables (ON USD, pata MEP) a un Excel
para compartir con un colega.

Toma los precios/rendimientos EN VIVO del server corriendo (panel `obligaciones_
negociables`, filtro MEP) y los enriquece con los términos del catálogo SQLite
(cupón, frecuencia, base, ley, serie/clase, tipo, vto). Escribe en `exports/`.

Uso:  py -3.12 scripts/export_ons_excel.py            (server en localhost:8000)
"""
from __future__ import annotations
import os
import re
import sys
import html
import json
import sqlite3
from datetime import date, datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config.settings import settings

BASE = os.environ.get("MONITOR_URL", "http://localhost:8000")
CCY = "MEP"
PROJ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# --- parsers de los valores formateados del panel --------------------------- #
def f_float(s):
    if s is None: return None
    if isinstance(s, (int, float)): return float(s)
    s = str(s).strip().replace(",", "")
    if not s or s in ("—", "-"): return None
    try: return float(s)
    except ValueError: return None

def f_pct(s):                     # "103.33%" / "-28.67%" / "+0.04%" → fracción
    s = (s or "").strip().replace("%", "").replace("+", "").replace(",", "")
    if not s or s in ("—", "-"): return None
    try: return float(s) / 100.0
    except ValueError: return None

def f_vol(s):                     # "35.1K" / "1.9B" / "0" → número
    s = (s or "").strip().replace(",", "")
    if not s or s in ("—", "-"): return None
    m = re.match(r"^([\d.]+)\s*([KMB]?)$", s, re.I)
    if not m: return f_float(s)
    return float(m.group(1)) * {"": 1, "K": 1e3, "M": 1e6, "B": 1e9}[m.group(2).upper()]

def f_date(iso, fallback):        # ISO del catálogo o dd/mm/yy del panel
    for s, fmt in ((iso, "%Y-%m-%d"), (fallback, "%d/%m/%y")):
        if s:
            try: return datetime.strptime(s.strip()[:10], fmt).date()
            except (ValueError, TypeError): pass
    return fallback

# --- 1) filas del panel en vivo --------------------------------------------- #
url = f"{BASE}/panels/obligaciones_negociables/rows?ccy={CCY}"
htmlrows = httpx.get(url, timeout=15).text
panel = []
# El endpoint /rows devuelve TODAS las patas (ARS/MEP/CABLE); el filtro de moneda
# es client-side (data-ccy). Nos quedamos solo con la pata MEP (…D, precio en USD).
for attrs, inner in re.findall(r"<tr([^>]*)>(.*?)</tr>", htmlrows, re.S):
    if "data-ticker" not in attrs:
        continue
    ccy_m = re.search(r'data-ccy="([^"]+)"', attrs)
    if not ccy_m or ccy_m.group(1) != CCY:
        continue
    tds = re.findall(r"<td[^>]*>(.*?)</td>", inner, re.S)
    txt = [html.unescape(re.sub(r"<[^>]+>", "", td)).strip() for td in tds]
    if len(txt) < 10:  # fila incompleta / "Cargando…"
        continue
    panel.append({
        "ticker": txt[0], "emisor": txt[1], "vto_panel": txt[2],
        "precio": f_float(txt[3]), "vtec": f_float(txt[4]), "paridad": f_pct(txt[5]),
        "tir": f_pct(txt[6]), "md": f_float(txt[7]), "var_dia": f_pct(txt[8]),
        "vol": f_vol(txt[9]),
    })

# --- 2) enriquecido del catálogo (por cualquier ticker de la fila-bono) ------ #
con = sqlite3.connect(str(settings.catalog_db))
enrich = {}
for tk, mep, ccl, rf in con.execute(
    "SELECT ticker, ticker_mep, ticker_ccl, raw_fields FROM instruments "
    "WHERE sheet='Obligaciones_Negociables'"):
    d = json.loads(rf) if rf else {}
    info = {
        "serie": d.get("serie_clase") or "",
        "ley": d.get("ley_aplicable") or "",
        "tipo": d.get("tipo") or "",
        "cupon": f_float(d.get("cupon anual %")),
        "freq": f_float(d.get("frecuencia pagos")),
        "base": d.get("base calculo") or "",
        "vto_iso": d.get("fecha_vencimiento") or "",
    }
    for t in (tk, mep, ccl):
        if t: enrich[t.upper()] = info
con.close()

# --- 3) join + filas finales ------------------------------------------------ #
COLS = [
    ("Ticker", "ticker", "text", 11),
    ("Emisor", "emisor", "text", 34),
    ("Serie / Clase", "serie", "text", 16),
    ("Ley", "ley", "text", 11),
    ("Tipo", "tipo", "text", 13),
    ("Vto", "vto", "date", 12),
    ("Cupón", "cupon", "pct", 9),
    ("Frec/año", "freq", "int", 9),
    ("Base", "base", "text", 10),
    ("Precio (USD)", "precio", "num", 12),
    ("V. Técnico", "vtec", "num", 12),
    ("Paridad", "paridad", "pct", 10),
    ("TIR", "tir", "pct", 10),
    ("MD", "md", "num", 8),
    ("Var día", "var_dia", "pct", 9),
    ("Volumen ($)", "vol", "vol", 14),
]
rows = []
for p in panel:
    e = enrich.get(p["ticker"].upper(), {})
    rows.append({
        **p,
        "serie": e.get("serie", ""), "ley": e.get("ley", ""), "tipo": e.get("tipo", ""),
        "cupon": (e.get("cupon") / 100.0) if e.get("cupon") is not None else None,
        "freq": e.get("freq"), "base": e.get("base", ""),
        "vto": f_date(e.get("vto_iso", ""), p["vto_panel"]),
    })
rows.sort(key=lambda r: (r["md"] is None, r["md"] or 0))   # por duración, como el panel

# --- 4) Excel con formato --------------------------------------------------- #
wb = Workbook(); ws = wb.active; ws.title = "ON USD"
navy = PatternFill("solid", fgColor="1C3170"); white = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="D2D8E6"); border = Border(bottom=thin)
gen = date.today().isoformat()

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
ws.cell(1, 1, "Obligaciones Negociables — ON USD (cotización pata MEP)").font = Font(bold=True, size=14)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLS))
ws.cell(2, 1, f"Generado {gen} · precios Data912 · {len(rows)} especies · TIR/Paridad/MD del motor del monitor").font = Font(italic=True, size=10, color="666666")

HEAD = 4
for j, (label, _key, _kind, width) in enumerate(COLS, start=1):
    c = ws.cell(HEAD, j, label); c.fill = navy; c.font = white
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(j)].width = width

FMT = {"num": "0.00", "pct": "0.00%", "vol": "#,##0", "int": "0", "date": "dd/mm/yyyy"}
for i, r in enumerate(rows, start=HEAD + 1):
    for j, (_label, key, kind, _w) in enumerate(COLS, start=1):
        c = ws.cell(i, j, r.get(key))
        c.border = border
        if kind in FMT and r.get(key) is not None:
            c.number_format = FMT[kind]
        c.alignment = Alignment(horizontal="left" if kind == "text" else "center")

ws.freeze_panes = f"A{HEAD + 1}"
ws.auto_filter.ref = f"A{HEAD}:{get_column_letter(len(COLS))}{HEAD + len(rows)}"

out_dir = os.path.join(PROJ, "exports"); os.makedirs(out_dir, exist_ok=True)
out = os.path.join(out_dir, f"Obligaciones_Negociables_ON_USD_{gen}.xlsx")
wb.save(out)
print(f"OK - {len(rows)} ONs -> {out}")
